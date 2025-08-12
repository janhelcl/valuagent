import io
import json
import os
from typing import Dict, Any

from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from pydantic import ValidationError
from dotenv import load_dotenv

from src import prompts, schemas, utils

load_dotenv()

app = FastAPI(title="Valuagent API", version="0.1.0")


def _pick_prompt(statement_type: str) -> str:
    if statement_type == "rozvaha":
        return prompts.balance_sheet_ocr_instructions
    if statement_type in {"vzz", "ziskaztrata", "vzzcz"}:  # accept a few aliases
        return prompts.profit_and_loss_ocr_instructions
    raise HTTPException(status_code=400, detail="Unsupported statement_type. Use 'rozvaha' or 'vzz'.")


def _validate_payload(statement_type: str, data_dict: Dict[str, Any], tolerance: int):
    if statement_type == "rozvaha":
        return schemas.BalanceSheet.model_validate_with_tolerance(data_dict, tolerance=tolerance)
    return schemas.ProfitAndLoss.model_validate_with_tolerance(data_dict, tolerance=tolerance)


def _export_excel(statement_type: str, model) -> io.BytesIO:
    # Lazy import so environments without Excel libs can still import API module
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    # Header
    if statement_type == "rozvaha":
        headers = ["Označení", "Brutto", "Korekce", "Netto", "Netto (minulé)"]
    else:
        headers = ["Označení", "Současné", "Minulé"]
    for idx, name in enumerate(headers, start=1):
        sheet.cell(row=1, column=idx, value=name)

    # Rows sorted by key
    sorted_rows = sorted(model.data.items(), key=lambda kv: kv[0])
    for r_idx, (row_id, row_obj) in enumerate(sorted_rows, start=2):
        sheet.cell(row=r_idx, column=1, value=row_id)
        if statement_type == "rozvaha":
            sheet.cell(row=r_idx, column=2, value=getattr(row_obj, "brutto", None))
            sheet.cell(row=r_idx, column=3, value=getattr(row_obj, "korekce", None))
            sheet.cell(row=r_idx, column=4, value=getattr(row_obj, "netto", 0))
            sheet.cell(row=r_idx, column=5, value=getattr(row_obj, "netto_minule", 0))
        else:
            sheet.cell(row=r_idx, column=2, value=getattr(row_obj, "současné", 0))
            sheet.cell(row=r_idx, column=3, value=getattr(row_obj, "minulé", 0))

    # Add validation report in a second sheet
    report_sheet = workbook.create_sheet("Report")
    report_text = model.summary_report()
    for i, line in enumerate(report_text.splitlines(), start=1):
        report_sheet.cell(row=i, column=1, value=line)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# Optional simple HTML form for manual testing
INDEX_HTML = """
<!DOCTYPE html>
<html>
  <head>
    <meta charset=\"utf-8\" />
    <title>Valuagent</title>
    <style>
      body { font-family: system-ui, Arial, sans-serif; margin: 2rem; }
      form { display: grid; gap: .75rem; max-width: 560px; }
      input, select, button { padding: .6rem; font-size: 1rem; }
    </style>
  </head>
  <body>
    <h2>Valuagent – Extract and Validate</h2>
    <form id=\"f\" action=\"/process\" method=\"post\" enctype=\"multipart/form-data\">
      <label>PDF <input type=\"file\" name=\"pdf\" accept=\"application/pdf\" required /></label>
      <label>Statement type
        <select name=\"statement_type\" required>
          <option value=\"rozvaha\">Rozvaha</option>
          <option value=\"vzz\">Výkaz zisku a ztráty</option>
        </select>
      </label>
      <label>Year <input type=\"number\" name=\"year\" value=\"2024\" required /></label>
      <label>Tolerance <input type=\"number\" name=\"tolerance\" value=\"0\" /></label>
      <button type=\"submit\">Process</button>
    </form>
    <p>Response will download as Excel if successful.</p>
  </body>
 </html>
"""


@app.get("/", response_class=HTMLResponse)
def index():
    return HTMLResponse(INDEX_HTML)


@app.post("/process")
async def process_pdf(
    pdf: UploadFile = File(...),
    statement_type: str = Form(...),  # "rozvaha" | "vzz"
    year: int = Form(...),
    tolerance: int = Form(0),
    return_json: bool = Form(False),
):
    # 1) Read uploaded file
    pdf_bytes = await pdf.read()
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    # 2) Prepare prompt
    system_prompt = _pick_prompt(statement_type)

    # 3) Call LLM via google-genai
    # We do a light integration; adjust model and safety settings as needed.
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Missing GOOGLE_API_KEY")

    try:
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=api_key)

        # Mirror notebook pattern: pass PDF bytes directly using Part.from_bytes
        result = client.models.generate_content(
            model=os.getenv("GENAI_MODEL", "gemini-2.5-pro"),
            contents=[
                types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
                system_prompt,
            ],
            config={
                "response_mime_type": "application/json",
            },
        )

        # Extract text response (JSON expected)
        text_response = result.text or ""
        if not text_response:
            raise RuntimeError("Empty response from model")

        # Robust JSON load
        try:
            data_dict = json.loads(text_response)
        except json.JSONDecodeError:
            data_dict = utils.load_json_from_text(text_response)

        # Ensure year in payload
        data_dict.setdefault("rok", year)

        # 4) Validate into pydantic model
        model_obj = _validate_payload(statement_type, data_dict, tolerance)

    except ValidationError as ve:
        raise HTTPException(status_code=422, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing failed: {e}")

    if return_json:
        return JSONResponse(model_obj.model_dump())

    # 5) Export to Excel
    excel_buffer = _export_excel(statement_type, model_obj)
    filename = f"valuagent_{statement_type}_{year}.xlsx"
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


