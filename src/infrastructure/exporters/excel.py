import io
import json
from pathlib import Path


def _load_row_index_map(statement_type: str) -> dict[int, str]:
    """Load a flat mapping of row number -> row name from resources.

    The JSON files are nested (with optional "sub_rows"), so we recursively
    flatten them into a simple dictionary for quick lookup when exporting.
    """
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    filename = (
        "balance_sheet_index.json" if statement_type == "rozvaha" else "profit_and_loss_index.json"
    )
    resource_path = resources_dir / filename

    try:
        with resource_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        return {}

    flat: dict[int, str] = {}

    def walk(node: dict):
        for key_str, value in node.items():
            try:
                key_int = int(key_str)
            except ValueError:
                # Skip non-integer keys defensively
                continue
            name = value.get("name")
            if isinstance(name, str):
                flat[key_int] = name
            sub = value.get("sub_rows")
            if isinstance(sub, dict):
                walk(sub)

    walk(data)
    return flat


def export_excel(statement_type: str, model) -> io.BytesIO:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add a dedicated column for human-readable row names using indices mapping
    headers = (
        ["Označení", "Název", "Brutto", "Korekce", "Netto", "Netto (minulé)"]
        if statement_type == "rozvaha"
        else ["Označení", "Název", "Současné", "Minulé"]
    )
    for idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=name)

    row_names = _load_row_index_map(statement_type)

    for r_idx, (row_id, row_obj) in enumerate(sorted(model.data.items(), key=lambda kv: kv[0]), start=2):
        # Column 1: numeric code, Column 2: human-readable name (if known)
        ws.cell(row=r_idx, column=1, value=row_id)
        ws.cell(row=r_idx, column=2, value=row_names.get(int(row_id)))
        if statement_type == "rozvaha":
            ws.cell(row=r_idx, column=3, value=getattr(row_obj, "brutto", None))
            ws.cell(row=r_idx, column=4, value=getattr(row_obj, "korekce", None))
            ws.cell(row=r_idx, column=5, value=getattr(row_obj, "netto", 0))
            ws.cell(row=r_idx, column=6, value=getattr(row_obj, "netto_minule", 0))
        else:
            ws.cell(row=r_idx, column=3, value=getattr(row_obj, "současné", 0))
            ws.cell(row=r_idx, column=4, value=getattr(row_obj, "minulé", 0))

    report = wb.create_sheet("Report")
    report_text = model.summary_report()
    for i, line in enumerate(report_text.splitlines(), start=1):
        report.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


