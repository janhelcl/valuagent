import io
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl

logger = logging.getLogger(__name__)


def load_predmet_oceneni_mapping() -> Dict[str, Dict[str, str]]:
    """Load the cell mapping for Předmět ocenění sheet."""
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    mapping_path = resources_dir / "predmet_oceneni_mapping.json"
    
    try:
        with mapping_path.open("r", encoding="utf-8") as f:
            mapping = json.load(f)
        logger.info(f"Loaded predmet_oceneni_mapping.json with {len(mapping)} row mappings")
        return mapping
    except FileNotFoundError:
        logger.error(f"Mapping file not found: {mapping_path}")
        return {}
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in mapping file: {e}")
        return {}


def load_rozvaha_mapping() -> Dict[str, Dict[str, str]]:
    """Load the row mapping for Rozvaha sheet."""
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    mapping_path = resources_dir / "rozvaha_mapping.json"
    
    try:
        with mapping_path.open("r", encoding="utf-8") as f:
            mapping = json.load(f)
        logger.info(f"Loaded rozvaha_mapping.json with {len(mapping)} row mappings")
        return mapping
    except FileNotFoundError:
        logger.error(f"Mapping file not found: {mapping_path}")
        return {}
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in mapping file: {e}")
        return {}


def load_vysledovka_mapping() -> Dict[str, Dict[str, str]]:
    """Load the row mapping for Výsledovka sheet."""
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    mapping_path = resources_dir / "vysledovka_mapping.json"
    
    try:
        with mapping_path.open("r", encoding="utf-8") as f:
            mapping = json.load(f)
        logger.info(f"Loaded vysledovka_mapping.json with {len(mapping)} row mappings")
        return mapping
    except FileNotFoundError:
        logger.error(f"Mapping file not found: {mapping_path}")
        return {}
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in mapping file: {e}")
        return {}


def find_latest_balance_sheet(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Find the balance sheet result with the latest year."""
    balance_sheets = [r for r in results if r["statement_type"] == "rozvaha"]
    
    if not balance_sheets:
        raise ValueError("No balance sheet found in results")
    
    # Sort by year (rok) in descending order
    balance_sheets.sort(key=lambda x: getattr(x["model"], "rok", 0), reverse=True)
    latest = balance_sheets[0]
    
    latest_year = getattr(latest["model"], "rok", "unknown")
    logger.info(f"Using latest balance sheet from year {latest_year} (from file: {latest['original']})")
    
    return latest


def get_sorted_balance_sheets(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Get all balance sheets sorted by year (newest first)."""
    balance_sheets = [r for r in results if r["statement_type"] == "rozvaha"]
    
    if not balance_sheets:
        return []
    
    # Sort by year (rok) in descending order (newest first)
    balance_sheets.sort(key=lambda x: getattr(x["model"], "rok", 0), reverse=True)
    
    years = [getattr(bs["model"], "rok", "unknown") for bs in balance_sheets]
    logger.info(f"Found {len(balance_sheets)} balance sheets for years: {years}")
    
    return balance_sheets


def get_sorted_profit_loss_statements(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Get all profit and loss statements sorted by year (newest first)."""
    profit_loss = [r for r in results if r["statement_type"] == "vzz"]
    
    if not profit_loss:
        return []
    
    # Sort by year (rok) in descending order (newest first)
    profit_loss.sort(key=lambda x: getattr(x["model"], "rok", 0), reverse=True)
    
    years = [getattr(pl["model"], "rok", "unknown") for pl in profit_loss]
    logger.info(f"Found {len(profit_loss)} profit and loss statements for years: {years}")
    
    return profit_loss


def fill_predmet_oceneni_sheet(workbook: openpyxl.Workbook, balance_sheet_result: Dict[str, Any], disambiguation_info: Dict[str, Any] = None) -> None:
    """Fill the Předmět ocenění sheet with balance sheet data."""
    sheet_name = "Předmět oce"
    
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {workbook.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in template")
    
    sheet = workbook[sheet_name]
    mapping = load_predmet_oceneni_mapping()
    
    if not mapping:
        logger.warning("No mapping data available, skipping sheet fill")
        return
    
    # Fill date cells E2, F2, G2, H2 with disambiguation date
    if disambiguation_info and disambiguation_info.get("datum"):
        date_str = disambiguation_info["datum"]
        try:
            # Parse the date (expected format: YYYY-MM-DD)
            from datetime import datetime
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d.%m.%Y")
            
            # Fill all date cells
            date_cells = ["E2", "F2", "G2", "H2"]
            for cell in date_cells:
                sheet[cell] = formatted_date
                logger.debug(f"Set date cell {cell} = {formatted_date}")
            
            logger.info(f"Filled date cells with {formatted_date}")
        except ValueError as e:
            logger.warning(f"Could not parse date '{date_str}': {e}")
    else:
        logger.debug("No disambiguation date available for Předmět ocenění sheet")
    
    balance_sheet = balance_sheet_result.get("model")
    # Support fallback to raw dict if model is not present
    if balance_sheet is not None:
        balance_data = getattr(balance_sheet, "data", {})
    else:
        raw = balance_sheet_result.get("raw") or {}
        balance_data = {}
        raw_data = raw.get("data") or {}
        # Normalize string keys to int where possible
        for k, v in raw_data.items():
            try:
                rid = int(k)
                balance_data[rid] = v
            except Exception:
                continue
    
    filled_count = 0
    missing_count = 0
    
    logger.info(f"Filling {sheet_name} sheet with {len(balance_data)} balance sheet rows")
    
    for row_id, row_data in balance_data.items():
        row_id_str = str(row_id)
        
        if row_id_str not in mapping:
            logger.debug(f"Row ID {row_id} not found in mapping, skipping")
            missing_count += 1
            continue
        
        cell_mapping = mapping[row_id_str]
        row_name = cell_mapping.get("name", f"Row {row_id}")
        
        try:
            # Access via model attributes or raw dict
            def get_attr(obj, name: str) -> Optional[int]:
                if hasattr(obj, name):
                    return getattr(obj, name)
                if isinstance(obj, dict):
                    return obj.get(name)
                return None

            # Fill brutto
            if "brutto" in cell_mapping:
                val = get_attr(row_data, "brutto")
                if val is not None:
                    sheet[cell_mapping["brutto"]] = val
                    logger.debug(f"Set {cell_mapping['brutto']} = {val} (brutto for {row_name})")
            # Fill korekce
            if "korekce" in cell_mapping:
                val = get_attr(row_data, "korekce")
                if val is not None:
                    sheet[cell_mapping["korekce"]] = val
                    logger.debug(f"Set {cell_mapping['korekce']} = {val} (korekce for {row_name})")
            # Fill netto
            if "netto" in cell_mapping:
                val = get_attr(row_data, "netto")
                if val is not None:
                    sheet[cell_mapping["netto"]] = val
                    logger.debug(f"Set {cell_mapping['netto']} = {val} (netto for {row_name})")

            filled_count += 1

        except Exception as e:
            logger.error(f"Error filling row {row_id} ({row_name}): {e}")
            continue
    
    logger.info(f"Successfully filled {filled_count} rows in {sheet_name} sheet, {missing_count} rows not mapped")


def fill_rozvaha_sheet(workbook: openpyxl.Workbook, balance_sheet_results: List[Dict[str, Any]]) -> None:
    """Fill the Rozvaha sheet with multiple years of balance sheet data."""
    sheet_name = "Rozvaha"
    
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {workbook.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in template")
    
    sheet = workbook[sheet_name]
    mapping = load_rozvaha_mapping()
    
    if not mapping:
        logger.warning("No Rozvaha mapping data available, skipping sheet fill")
        return
    
    if not balance_sheet_results:
        logger.warning("No balance sheet results available for Rozvaha sheet")
        return
    
    # Fill latest year in J3 (latest year from Předmět ocenění)
    latest_year = getattr(balance_sheet_results[0]["model"], "rok", "")
    if latest_year:
        sheet["J3"] = latest_year
        logger.debug(f"Set latest year header J3 = {latest_year}")
    
    # Column mapping: I = 2nd latest, H = 3rd latest, G = 4th latest, F = 5th latest
    year_columns = ['I', 'H', 'G', 'F']
    max_years = len(year_columns)
    
    # Collect all possible years with their data sources
    all_years = set()
    year_to_source = {}
    
    # First pass: collect all netto years (skip latest as it's in Předmět ocenění)
    for bs in balance_sheet_results[1:]:  # Skip latest year
        model = bs.get("model")
        if model is not None:
            year = getattr(model, "rok", 0)
            if year > 0:
                all_years.add(year)
                year_to_source[year] = (bs, 'netto')
    
    # Second pass: add netto_minule years if we have space
    for bs in balance_sheet_results:
        model = bs.get("model")
        if model is not None:
            year = getattr(model, "rok", 0) - 1  # netto_minule represents previous year
            if year > 0 and year not in all_years:
                all_years.add(year)
                year_to_source[year] = (bs, 'netto_minule')
    
    # Sort years descending and take only what fits in columns
    sorted_years = sorted(all_years, reverse=True)[:max_years]
    
    # Build data sources list: (balance_sheet_result, year, data_source)
    data_sources = []
    for year in sorted_years:
        if year in year_to_source:
            bs, source = year_to_source[year]
            data_sources.append((bs, year, source))
    
    if not data_sources:
        logger.info("No historical balance sheet data available for Rozvaha sheet")
        return
    
    logger.info(f"Filling Rozvaha sheet with {len(data_sources)} years of historical data")
    
    # Fill year headers in row 3
    for i, (balance_sheet_result, year, data_source) in enumerate(data_sources):
        column = year_columns[i]
        year_cell = f"{column}3"
        sheet[year_cell] = year
        logger.debug(f"Set year header {year_cell} = {year} (from {data_source})")
    
    # Fill data for each year
    total_filled = 0
    total_missing = 0
    
    for year_idx, (balance_sheet_result, year, data_source) in enumerate(data_sources):
        column = year_columns[year_idx]
        balance_sheet = balance_sheet_result.get("model")
        if balance_sheet is not None:
            balance_data = getattr(balance_sheet, "data", {})
        else:
            raw = balance_sheet_result.get("raw") or {}
            raw_data = raw.get("data") or {}
            balance_data = {}
            for k, v in raw_data.items():
                try:
                    rid = int(k)
                    balance_data[rid] = v
                except Exception:
                    continue
        
        filled_count = 0
        missing_count = 0
        
        logger.debug(f"Filling column {column} with year {year} data from {data_source} ({len(balance_data)} rows)")
        
        for row_id, row_data in balance_data.items():
            row_id_str = str(row_id)
            
            if row_id_str not in mapping:
                logger.debug(f"Row ID {row_id} not found in Rozvaha mapping, skipping")
                missing_count += 1
                continue
            
            cell_mapping = mapping[row_id_str]
            row_name = cell_mapping.get("name", f"Row {row_id}")
            excel_row = cell_mapping.get("netto")
            
            if not excel_row:
                logger.warning(f"No row mapping found for row ID {row_id}")
                missing_count += 1
                continue
            
            try:
                # Get the value based on data source
                def get_val(obj, name: str) -> Optional[int]:
                    if hasattr(obj, name):
                        return getattr(obj, name)
                    if isinstance(obj, dict):
                        return obj.get(name)
                    return None

                value = None
                if data_source == 'netto':
                    value = get_val(row_data, 'netto')
                elif data_source == 'netto_minule':
                    value = get_val(row_data, 'netto_minule')

                if value is not None:
                    cell_address = f"{column}{excel_row}"

                    # Check if cell contains a formula - if so, skip it to preserve template logic
                    existing_cell = sheet[cell_address]
                    if existing_cell.data_type == 'f':  # 'f' means formula
                        logger.debug(f"Skipping {cell_address} - contains formula: {existing_cell.value}")
                        missing_count += 1
                    else:
                        sheet[cell_address] = value
                        logger.debug(f"Set {cell_address} = {value} ({data_source} for {row_name})")
                        filled_count += 1
                else:
                    logger.debug(f"No {data_source} value for row {row_id} ({row_name})")
                    missing_count += 1

            except Exception as e:
                logger.error(f"Error filling row {row_id} ({row_name}) in column {column}: {e}")
                missing_count += 1
                continue
        
        logger.info(f"Column {column} (year {year} from {data_source}): filled {filled_count} rows, {missing_count} missing")
        total_filled += filled_count
        total_missing += missing_count
    
    logger.info(f"Successfully filled Rozvaha sheet: {total_filled} total values, {total_missing} total missing")


def fill_vysledovka_sheet(workbook: openpyxl.Workbook, profit_loss_results: List[Dict[str, Any]]) -> None:
    """Fill the Výsledovka sheet with multiple years of profit and loss data."""
    sheet_name = "Výsledovka"
    
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {workbook.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in template")
    
    sheet = workbook[sheet_name]
    mapping = load_vysledovka_mapping()
    
    if not mapping:
        logger.warning("No Výsledovka mapping data available, skipping sheet fill")
        return
    
    if not profit_loss_results:
        logger.warning("No profit and loss results available for Výsledovka sheet")
        return
    
    # Column mapping: J = latest, I = 2nd latest, H = 3rd latest, G = 4th latest, F = 5th latest
    year_columns = ['J', 'I', 'H', 'G', 'F']
    max_years = len(year_columns)
    
    # Collect all possible years with their data sources
    all_years = set()
    year_to_source = {}
    
    # First pass: collect all současné years
    for pl in profit_loss_results:
        model = pl.get("model")
        if model is not None:
            year = getattr(model, "rok", 0)
            if year > 0:
                all_years.add(year)
                year_to_source[year] = (pl, 'současné')
    
    # Second pass: add minulé years if not already present
    for pl in profit_loss_results:
        model = pl.get("model")
        if model is not None:
            year = getattr(model, "rok", 0) - 1  # minulé represents previous year
            if year > 0 and year not in all_years:
                all_years.add(year)
                year_to_source[year] = (pl, 'minulé')
    
    # Sort years descending and take only what fits in columns
    sorted_years = sorted(all_years, reverse=True)[:max_years]
    
    # Build data sources list: (profit_loss_result, year, data_source)
    data_sources = []
    for year in sorted_years:
        if year in year_to_source:
            pl, source = year_to_source[year]
            data_sources.append((pl, year, source))
    
    if not data_sources:
        logger.info("No profit and loss data available for Výsledovka sheet")
        return
    
    logger.info(f"Filling Výsledovka sheet with {len(data_sources)} years of data")
    
    # Fill data for each year
    total_filled = 0
    total_missing = 0
    
    for year_idx, (profit_loss_result, year, data_source) in enumerate(data_sources):
        column = year_columns[year_idx]
        profit_loss = profit_loss_result.get("model")
        if profit_loss is not None:
            profit_loss_data = getattr(profit_loss, "data", {})
        else:
            raw = profit_loss_result.get("raw") or {}
            raw_data = raw.get("data") or {}
            profit_loss_data = {}
            for k, v in raw_data.items():
                try:
                    rid = int(k)
                    profit_loss_data[rid] = v
                except Exception:
                    continue
        
        filled_count = 0
        missing_count = 0
        
        logger.debug(f"Filling column {column} with year {year} data from {data_source} ({len(profit_loss_data)} rows)")
        
        for row_id, row_data in profit_loss_data.items():
            row_id_str = str(row_id)
            
            if row_id_str not in mapping:
                logger.debug(f"Row ID {row_id} not found in Výsledovka mapping, skipping")
                missing_count += 1
                continue
            
            cell_mapping = mapping[row_id_str]
            row_name = cell_mapping.get("name", f"Row {row_id}")
            excel_row = cell_mapping.get("netto")
            
            if not excel_row:
                logger.warning(f"No row mapping found for row ID {row_id}")
                missing_count += 1
                continue
            
            try:
                # Get the value based on data source
                def get_val(obj, name: str) -> Optional[int]:
                    if hasattr(obj, name):
                        return getattr(obj, name)
                    if isinstance(obj, dict):
                        return obj.get(name)
                    return None

                value = None
                if data_source == 'současné':
                    value = get_val(row_data, 'současné')
                elif data_source == 'minulé':
                    value = get_val(row_data, 'minulé')

                if value is not None:
                    cell_address = f"{column}{excel_row}"

                    # Check if cell contains a formula - if so, skip it to preserve template logic
                    existing_cell = sheet[cell_address]
                    if existing_cell.data_type == 'f':  # 'f' means formula
                        logger.debug(f"Skipping {cell_address} - contains formula: {existing_cell.value}")
                        missing_count += 1
                    else:
                        sheet[cell_address] = value
                        logger.debug(f"Set {cell_address} = {value} ({data_source} for {row_name})")
                        filled_count += 1
                else:
                    logger.debug(f"No {data_source} value for row {row_id} ({row_name})")
                    missing_count += 1

            except Exception as e:
                logger.error(f"Error filling row {row_id} ({row_name}) in column {column}: {e}")
                missing_count += 1
                continue
        
        logger.info(f"Column {column} (year {year} from {data_source}): filled {filled_count} rows, {missing_count} missing")
        total_filled += filled_count
        total_missing += missing_count
    
    logger.info(f"Successfully filled Výsledovka sheet: {total_filled} total values, {total_missing} total missing")


def add_data_quality_report(workbook: openpyxl.Workbook, results: List[Dict[str, Any]], inter_issues: List[str], tolerance: int) -> None:
    """Create a clear, Czech data quality report sheet summarizing retries, status, and issues."""
    sheet_name = "Kvalita dat"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Clear existing content (optional)
        for row in sheet[1:sheet.max_row]:
            for cell in row:
                cell.value = None
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Header
    sheet["A1"] = "Kvalita dat"
    sheet["A2"] = f"Tolerance: {tolerance}"

    # Overview table
    headers = ["Soubor", "Výkaz", "Rok", "Pokusy OCR", "Status", "Počet problémů"]
    for col, h in enumerate(headers, start=1):
        sheet.cell(row=4, column=col, value=h)

    row = 5
    for r in results:
        file_name = r.get("original")
        st = r.get("statement_type")
        model = r.get("model")
        rok = getattr(model, "rok", None) if model is not None else (r.get("raw") or {}).get("rok")
        attempts = r.get("ocr_attempts", 1)
        status = r.get("status", "ok")
        err_count = len(r.get("validation_errors") or [])
        values = [file_name, st, rok, attempts, "OK" if status == "ok" else "Chyby", err_count]
        for col, v in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=v)
        row += 1

    # Statement-level problems
    start = row + 1
    sheet.cell(row=start, column=1, value="Problémy ve výkazech (po opakování OCR)")
    row = start + 1
    for r in results:
        errs = r.get("validation_errors") or []
        if not errs:
            continue
        file_name = r.get("original")
        st = r.get("statement_type")
        model = r.get("model")
        rok = getattr(model, "rok", None) if model is not None else (r.get("raw") or {}).get("rok")
        sheet.cell(row=row, column=1, value=f"Soubor: {file_name}")
        sheet.cell(row=row, column=2, value=f"Výkaz: {st}")
        sheet.cell(row=row, column=3, value=f"Rok: {rok}")
        row += 1
        for msg in errs:
            sheet.cell(row=row, column=2, value=msg)
            row += 1
        row += 1

    # Interstatement issues
    sheet.cell(row=row, column=1, value="Problémy mezi výkazy a mezi roky")
    row += 1
    for msg in inter_issues or []:
        sheet.cell(row=row, column=2, value=msg)
        row += 1


def export_dcf_template(results: List[Dict[str, Any]], disambiguation_info: Dict[str, Any] = None, tolerance: int = 1) -> io.BytesIO:
    """Export results to DCF template, filling Předmět ocenění, Rozvaha, Výsledovka a Kvalita dat."""
    logger.info(f"Creating DCF template export from {len(results)} results")
    
    # Load the template
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    template_path = resources_dir / "DCF.xlsx"
    
    if not template_path.exists():
        raise FileNotFoundError(f"DCF template not found: {template_path}")
    
    logger.info(f"Loading DCF template from {template_path}")
    workbook = openpyxl.load_workbook(template_path)
    logger.debug(f"Template loaded with sheets: {workbook.sheetnames}")
    
    # Get all balance sheets sorted by year (newest first)
    sorted_balance_sheets = get_sorted_balance_sheets(results)
    
    if not sorted_balance_sheets:
        logger.error("No balance sheet data found in results")
        raise ValueError("No balance sheet data found in results")
    
    # Fill Předmět ocenění sheet with latest year and disambiguation date
    latest_balance_sheet = sorted_balance_sheets[0]
    fill_predmet_oceneni_sheet(workbook, latest_balance_sheet, disambiguation_info)
    
    # Fill Rozvaha sheet with historical years (skip latest)
    if len(sorted_balance_sheets) > 1:
        fill_rozvaha_sheet(workbook, sorted_balance_sheets)
    else:
        logger.info("Only one balance sheet year available, skipping Rozvaha sheet historical data")
    
    # Fill Výsledovka sheet with profit and loss data
    sorted_profit_loss = get_sorted_profit_loss_statements(results)
    if sorted_profit_loss:
        fill_vysledovka_sheet(workbook, sorted_profit_loss)
    else:
        logger.info("No profit and loss data available, skipping Výsledovka sheet")
    
    # Add Data Quality report sheet
    try:
        from src.services.quality import validate_interstatement
        inter_issues = validate_interstatement(results, tolerance)
        add_data_quality_report(workbook, results, inter_issues, tolerance)
    except Exception as e:
        logger.error(f"Failed to add Data Quality report: {e}", exc_info=True)

    # Save to BytesIO buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    logger.info(f"DCF template export completed, buffer size: {buffer.getbuffer().nbytes/1024:.1f}KB")
    return buffer
