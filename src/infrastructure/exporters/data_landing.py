"""
Data Landing Exporter

This module fills financial statement data into a user-provided Excel template.
The user's Excel file contains pre-configured sheets with formulas and structure.
This exporter only fills in the raw data values and dates.

- Data - Rozvaha: Fill numerical values and dates
  * AKTIVA section: Fills Brutto, Korekce, Netto values
  * PASIVA section: Fills Netto values only
  * Dates are filled in dd.mm.yyyy format at the top of columns
  * Only data values are filled - structure is pre-filled by user

- Data - Report Kvality: Quality report with validation results

The user uploads their Excel template along with PDFs, and we fill in the data.
"""

import io
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def get_datum_for_sorting(result: Dict[str, Any]) -> str:
    """Get datum from disambiguation info for sorting (YYYY-MM-DD format).
    
    Args:
        result: Result dict containing disambiguation_info with datum field
        
    Returns:
        Date string in YYYY-MM-DD format, or empty string if not available
        Empty string sorts before any valid date
    """
    disambiguation_info = result.get("disambiguation_info", {})
    if not disambiguation_info:
        return ""
    
    datum = disambiguation_info.get("datum")
    if not datum:
        return ""
    
    # Validate the date format
    try:
        datetime.strptime(datum, "%Y-%m-%d")
        return datum
    except ValueError as e:
        logger.warning(f"Could not parse datum '{datum}': {e}")
        return ""


def extract_year_from_datum(result: Dict[str, Any]) -> int:
    """Extract year from disambiguation datum field.
    
    Args:
        result: Result dict containing disambiguation_info with datum field
        
    Returns:
        Year as integer, or 0 if not available
    """
    datum = get_datum_for_sorting(result)
    if not datum:
        return 0
    
    try:
        date_obj = datetime.strptime(datum, "%Y-%m-%d")
        return date_obj.year
    except ValueError:
        return 0


def format_datum_for_excel(result: Dict[str, Any]) -> str:
    """Format the datum from disambiguation info for Excel display (dd.mm.yyyy).
    
    Args:
        result: Result dict containing disambiguation_info with datum field
        
    Returns:
        Formatted date string (dd.mm.yyyy) or empty string if not available
    """
    disambiguation_info = result.get("disambiguation_info", {})
    if not disambiguation_info:
        file_name = result.get("original", "unknown")
        logger.warning(f"No disambiguation_info found in result for {file_name}")
        return ""
    
    datum = disambiguation_info.get("datum")
    if not datum:
        file_name = result.get("original", "unknown")
        logger.warning(f"No datum found in disambiguation_info for {file_name}")
        return ""
    
    try:
        # Parse YYYY-MM-DD format and convert to dd.mm.yyyy
        date_obj = datetime.strptime(datum, "%Y-%m-%d")
        formatted = date_obj.strftime("%d.%m.%Y")
        logger.debug(f"Formatted datum '{datum}' to '{formatted}' for {result.get('original', 'unknown')}")
        return formatted
    except ValueError as e:
        logger.warning(f"Could not parse datum '{datum}': {e}")
        return ""


def load_balance_sheet_structure() -> Dict[int, Dict[str, Any]]:
    """Load the balance sheet structure with row labels and descriptions."""
    resources_dir = Path(__file__).resolve().parent.parent / "resources"
    index_path = resources_dir / "balance_sheet_index.json"
    
    try:
        with index_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Flatten structure to get row_id -> name mapping
        flat = {}
        def walk(node: dict, parent_label: str = ""):
            for key_str, value in node.items():
                try:
                    key_int = int(key_str)
                except ValueError:
                    continue
                name = value.get("name")
                if isinstance(name, str):
                    flat[key_int] = {"name": name, "label": parent_label}
                sub = value.get("sub_rows")
                if isinstance(sub, dict):
                    walk(sub, parent_label)
        
        walk(data)
        return flat
    except FileNotFoundError:
        logger.error(f"Balance sheet index not found: {index_path}")
        return {}


def get_sorted_balance_sheets(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Get all balance sheets sorted by date (newest first)."""
    balance_sheets = [r for r in results if r["statement_type"] == "rozvaha"]
    
    if not balance_sheets:
        return []
    
    # Log each balance sheet before sorting
    for i, bs in enumerate(balance_sheets):
        file_name = bs.get("original", "unknown")
        datum_raw = bs.get("disambiguation_info", {}).get("datum", "N/A")
        logger.info(f"Balance sheet {i+1}: {file_name} - raw datum={datum_raw}")
    
    # Sort by full date from disambiguation datum (newest first)
    balance_sheets.sort(key=lambda x: get_datum_for_sorting(x), reverse=True)
    
    dates = [format_datum_for_excel(bs) or get_datum_for_sorting(bs) for bs in balance_sheets]
    logger.info(f"Found {len(balance_sheets)} balance sheets for dates: {dates}")
    
    return balance_sheets


def get_sorted_profit_loss_statements(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Get all P&L statements sorted by date (newest first)."""
    profit_loss = [r for r in results if r["statement_type"] == "vzz"]
    
    if not profit_loss:
        return []
    
    # Sort by full date from disambiguation datum (newest first)
    profit_loss.sort(key=lambda x: get_datum_for_sorting(x), reverse=True)
    
    dates = [format_datum_for_excel(pl) or get_datum_for_sorting(pl) for pl in profit_loss]
    logger.info(f"Found {len(profit_loss)} P&L statements for dates: {dates}")
    
    return profit_loss


def get_row_number_column() -> int:
    """Return the fixed column for row numbers (ř.) - always column D."""
    return 4  # Column D


def get_data_start_column() -> int:
    """Return the fixed column where data starts - always column E."""
    return 5  # Column E (after column D)


def fill_rozvaha_sheet(workbook: openpyxl.Workbook, balance_sheet_results: List[Dict[str, Any]], offset: int = 0) -> None:
    """Fill numerical data and dates into existing Data - Rozvaha sheet.
    
    Args:
        workbook: Excel workbook to fill
        balance_sheet_results: List of balance sheet results sorted by year
        offset: Number of years to skip from the left (default 0)
                Each year uses 3 columns (Brutto, Korekce, Netto)
    """
    sheet_name = "Data - Rozvaha"
    
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {workbook.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in uploaded template")
    
    sheet = workbook[sheet_name]
    
    if not balance_sheet_results:
        logger.warning("No balance sheet results available for Data - Rozvaha sheet")
        return
    
    # Limit to 7 years
    balance_sheet_results = balance_sheet_results[:7]
    
    # Fixed columns: D for row numbers, E for data start
    row_num_col = get_row_number_column()  # Column D
    base_data_col = get_data_start_column()  # Column E
    
    # Apply offset: each year in Rozvaha uses 3 columns (Brutto, Korekce, Netto)
    data_start_col = base_data_col + (offset * 3)
    
    logger.info(f"Filling {sheet_name} with {len(balance_sheet_results)} years starting at column {data_start_col} (offset={offset}, row numbers in column {row_num_col})")
    
    # Fill dates in row 1 (dd.mm.yyyy format from disambiguation datum)
    col_index = data_start_col
    result_to_columns = {}  # Map result to list of column indices
    
    for bs_result in balance_sheet_results:
        # Get date from disambiguation info
        date_str = format_datum_for_excel(bs_result)
        file_name = bs_result.get("original", "unknown")
        
        if not date_str:
            logger.warning(f"No date available for balance sheet from {file_name}, skipping")
            continue
        
        logger.info(f"Filling Rozvaha from {file_name} with date {date_str}")
        
        # Store column positions for this result
        result_columns = []
        
        # AKTIVA section: 3 columns per year (Brutto, Korekce, Netto)
        # Fill the same date for all 3 columns
        for i in range(3):
            sheet.cell(row=1, column=col_index, value=date_str)
            result_columns.append(col_index)
            col_index += 1
        
        result_to_columns[id(bs_result)] = result_columns
        logger.debug(f"Set dates for {date_str} in columns {result_columns}")
    
    # Create a mapping of row_number -> sheet_row for efficient lookup
    # Scan the row number column to find where each row ID is located
    row_id_to_sheet_row = {}
    for row_idx in range(3, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=row_num_col).value
        if cell_value is not None:
            try:
                # Convert to integer (handles both int and string representations)
                row_id = int(cell_value)
                if row_id > 0:
                    row_id_to_sheet_row[row_id] = row_idx
            except (ValueError, TypeError):
                pass
    
    logger.info(f"Found {len(row_id_to_sheet_row)} row mappings in template: {list(row_id_to_sheet_row.keys())[:10]}...")
    logger.info(f"Result to columns mapping: {len(result_to_columns)} results")
    
    # Fill data for each balance sheet result
    total_filled = 0
    for bs_result in balance_sheet_results:
        result_id = id(bs_result)
        if result_id not in result_to_columns:
            logger.warning(f"No columns allocated for result, skipping")
            continue
        
        columns = result_to_columns[result_id]
        date_str = format_datum_for_excel(bs_result)
        
        model = bs_result.get("model")
        if not model:
            logger.warning(f"No model found for result with date {date_str}")
            continue
        
        logger.info(f"Processing {date_str} with {len(model.data)} rows")
        
        # columns = [brutto_col, korekce_col, netto_col]
        brutto_col, korekce_col, netto_col = columns
        logger.info(f"Using columns for {date_str}: Brutto={brutto_col}, Korekce={korekce_col}, Netto={netto_col}")
        
        filled_count = 0
        skipped_count = 0
        for row_id, row_data in model.data.items():
            # Find which Excel row corresponds to this row_id
            if row_id not in row_id_to_sheet_row:
                logger.debug(f"Row ID {row_id} not found in template, skipping")
                skipped_count += 1
                continue
            
            excel_row = row_id_to_sheet_row[row_id]
            logger.debug(f"  Processing row_id={row_id} -> excel_row={excel_row}")
            
            # Determine if this is AKTIVA or PASIVA
            # AKTIVA (< 78): Fill Brutto, Korekce, Netto
            # PASIVA (>= 78): Fill only Netto (in the third column position)
            
            if row_id < 78:
                # AKTIVA section - fill all 3 columns
                # Brutto
                brutto_val = getattr(row_data, "brutto", None)
                if brutto_val is not None:
                    sheet.cell(row=excel_row, column=brutto_col, value=brutto_val)
                    logger.debug(f"  Filled row {row_id} Brutto at ({excel_row}, {brutto_col}): {brutto_val}")
                
                # Korekce (as negative)
                korekce_val = getattr(row_data, "korekce", None)
                if korekce_val is not None:
                    try:
                        korekce_val = -abs(int(korekce_val))
                    except Exception:
                        pass
                    sheet.cell(row=excel_row, column=korekce_col, value=korekce_val)
                    logger.debug(f"  Filled row {row_id} Korekce at ({excel_row}, {korekce_col}): {korekce_val}")
                
                # Netto
                netto_val = getattr(row_data, "netto", None)
                if netto_val is not None:
                    sheet.cell(row=excel_row, column=netto_col, value=netto_val)
                    logger.debug(f"  Filled row {row_id} Netto at ({excel_row}, {netto_col}): {netto_val}")
                
                filled_count += 1
            else:
                # PASIVA section - fill only Netto column
                netto_val = getattr(row_data, "netto", None)
                if netto_val is not None:
                    # For PASIVA, use the netto column position
                    sheet.cell(row=excel_row, column=netto_col, value=netto_val)
                    logger.debug(f"  Filled PASIVA row {row_id} Netto at ({excel_row}, {netto_col}): {netto_val}")
                filled_count += 1
        
        logger.info(f"{date_str}: Filled {filled_count} rows, Skipped {skipped_count} rows (not in template)")
        total_filled += filled_count
    
    logger.info(f"Successfully filled {sheet_name} sheet: {total_filled} total values")


def fill_vysledovka_sheet(workbook: openpyxl.Workbook, profit_loss_results: List[Dict[str, Any]], offset: int = 0) -> None:
    """Fill numerical data and dates into existing Data - Výsledovka sheet.
    
    Args:
        workbook: Excel workbook to fill
        profit_loss_results: List of P&L results sorted by year
        offset: Number of years to skip from the left (default 0)
                Each year uses 1 column
    """
    sheet_name = "Data - Výsledovka"
    
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {workbook.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in uploaded template")
    
    sheet = workbook[sheet_name]
    
    if not profit_loss_results:
        logger.warning("No P&L results available for Data - Výsledovka sheet")
        return
    
    # Limit to 7 years
    profit_loss_results = profit_loss_results[:7]
    
    # Fixed columns: D for row numbers, E for data start
    row_num_col = get_row_number_column()  # Column D
    base_data_col = get_data_start_column()  # Column E
    
    # Apply offset: each year in Výsledovka uses 1 column
    data_start_col = base_data_col + offset
    
    logger.info(f"Filling {sheet_name} with {len(profit_loss_results)} years starting at column {data_start_col} (offset={offset}, row numbers in column {row_num_col})")
    
    # Fill dates in row 1 (dd.mm.yyyy format from disambiguation datum)
    col_index = data_start_col
    result_to_column = {}  # Map result to column index (P&L has 1 column per year)
    
    for pl_result in profit_loss_results:
        # Get date from disambiguation info
        date_str = format_datum_for_excel(pl_result)
        file_name = pl_result.get("original", "unknown")
        
        if not date_str:
            logger.warning(f"No date available for P&L from {file_name}, skipping")
            continue
        
        logger.info(f"Filling Výsledovka from {file_name} with date {date_str}")
        
        # P&L: just 1 column per year (not 3 like balance sheet)
        sheet.cell(row=1, column=col_index, value=date_str)
        result_to_column[id(pl_result)] = col_index
        logger.debug(f"Set date for {date_str} in column {col_index}")
        
        col_index += 1
    
    # Create a mapping of row_number -> sheet_row for efficient lookup
    # Scan the row number column to find where each row ID is located
    # For P&L, data starts at row 2 (not row 3 like in Rozvaha)
    row_id_to_sheet_row = {}
    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 for P&L
        cell_value = sheet.cell(row=row_idx, column=row_num_col).value
        if cell_value is not None:
            try:
                # Convert to integer (handles both int and string representations)
                row_id = int(cell_value)
                if row_id > 0:
                    row_id_to_sheet_row[row_id] = row_idx
            except (ValueError, TypeError):
                pass
    
    logger.info(f"Found {len(row_id_to_sheet_row)} row mappings in template: {list(row_id_to_sheet_row.keys())[:10]}...")
    logger.info(f"Result to column mapping: {len(result_to_column)} results")
    
    # Fill data for each P&L result
    total_filled = 0
    for pl_result in profit_loss_results:
        result_id = id(pl_result)
        if result_id not in result_to_column:
            logger.warning(f"No column allocated for result, skipping")
            continue
        
        column = result_to_column[result_id]
        date_str = format_datum_for_excel(pl_result)
        
        model = pl_result.get("model")
        if not model:
            logger.warning(f"No model found for result with date {date_str}")
            continue
        
        logger.info(f"Processing {date_str} with {len(model.data)} rows")
        logger.info(f"Using column {column} for {date_str}")
        
        filled_count = 0
        skipped_count = 0
        for row_id, row_data in model.data.items():
            # Find which Excel row corresponds to this row_id
            if row_id not in row_id_to_sheet_row:
                logger.debug(f"Row ID {row_id} not found in template, skipping")
                skipped_count += 1
                continue
            
            excel_row = row_id_to_sheet_row[row_id]
            logger.debug(f"  Processing row_id={row_id} -> excel_row={excel_row}")
            
            # P&L uses 'současné' field for current year data
            value = getattr(row_data, "současné", None)
            if value is not None:
                sheet.cell(row=excel_row, column=column, value=value)
                logger.debug(f"  Filled row {row_id} at ({excel_row}, {column}): {value}")
                filled_count += 1
        
        logger.info(f"{date_str}: Filled {filled_count} rows, Skipped {skipped_count} rows (not in template)")
        total_filled += filled_count
    
    logger.info(f"Successfully filled {sheet_name} sheet: {total_filled} total values")


def fill_quality_report_sheet(workbook: openpyxl.Workbook, results: List[Dict[str, Any]], inter_issues: List[str], tolerance: int) -> None:
    """Fill or create the Data - Report Kvality sheet with quality information."""
    sheet_name = "Data - Report Kvality"
    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        logger.info(f"Found existing '{sheet_name}' sheet, clearing and filling with data")
        # Clear existing content
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
    else:
        logger.info(f"Creating new '{sheet_name}' sheet")
        sheet = workbook.create_sheet(sheet_name)
    
    # Header
    sheet["A1"] = "Kvalita dat"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Tolerance: {tolerance}"
    sheet["A3"] = "Poznámka: Všechny hodnoty jsou normalizovány na tisíce (tis. Kč)"
    sheet["A3"].font = Font(italic=True)
    
    # Overview table
    headers = ["Soubor", "Výkaz", "Datum", "Pokusy OCR", "Status", "Počet problémů"]
    for col, h in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)
    
    row = 6
    for r in results:
        file_name = r.get("original")
        st = r.get("statement_type")
        # Get date from disambiguation info, or fallback to rok from model
        date_str = format_datum_for_excel(r)
        if not date_str:
            model = r.get("model")
            rok = getattr(model, "rok", None) if model is not None else (r.get("raw") or {}).get("rok")
            date_display = rok if rok else "N/A"
        else:
            date_display = date_str
        attempts = r.get("ocr_attempts", 1)
        status = r.get("status", "ok")
        err_count = len(r.get("validation_errors") or [])
        values = [file_name, st, date_display, attempts, "OK" if status == "ok" else "Chyby", err_count]
        for col, v in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=v)
        row += 1
    
    # Statement-level problems
    start = row + 1
    sheet.cell(row=start, column=1, value="Problémy ve výkazech (po opakování OCR)").font = Font(bold=True)
    row = start + 1
    
    found_errors = False
    for r in results:
        errs = r.get("validation_errors") or []
        if not errs:
            continue
        found_errors = True
        file_name = r.get("original")
        st = r.get("statement_type")
        # Get date from disambiguation info, or fallback to rok from model
        date_str = format_datum_for_excel(r)
        if not date_str:
            model = r.get("model")
            rok = getattr(model, "rok", None) if model is not None else (r.get("raw") or {}).get("rok")
            date_display = rok if rok else "N/A"
        else:
            date_display = date_str
        sheet.cell(row=row, column=1, value=f"Soubor: {file_name}")
        sheet.cell(row=row, column=2, value=f"Výkaz: {st}")
        sheet.cell(row=row, column=3, value=f"Datum: {date_display}")
        row += 1
        for msg in errs:
            sheet.cell(row=row, column=2, value=msg)
            row += 1
        row += 1
    
    if not found_errors:
        sheet.cell(row=row, column=1, value="Žádné problémy")
        row += 2
    
    # Interstatement issues
    sheet.cell(row=row, column=1, value="Problémy mezi výkazy a mezi roky").font = Font(bold=True)
    row += 1
    if inter_issues:
        for msg in inter_issues:
            sheet.cell(row=row, column=1, value=msg)
            row += 1
    else:
        sheet.cell(row=row, column=1, value="Žádné problémy")
        row += 1
    
    logger.info(f"Successfully created {sheet_name} sheet")


def export_data_landing(results: List[Dict[str, Any]], template_bytes: bytes, tolerance: int = 1, offset: int = 0) -> io.BytesIO:
    """
    Fill financial data into user-provided Excel template.
    
    Args:
        results: Processed financial statement results
        template_bytes: User's Excel template file as bytes
        tolerance: Validation tolerance
        offset: Number of years to skip from the left (default 0)
                Useful when newest data isn't available yet
    
    Returns:
        BytesIO buffer with filled Excel file
    """
    logger.info(f"Filling data landing template with {len(results)} results (offset={offset})")
    
    # Load the user's template
    template_buffer = io.BytesIO(template_bytes)
    workbook = openpyxl.load_workbook(template_buffer)
    logger.info(f"Loaded template with sheets: {workbook.sheetnames}")
    
    # Get sorted balance sheets
    sorted_balance_sheets = get_sorted_balance_sheets(results)
    
    if sorted_balance_sheets:
        fill_rozvaha_sheet(workbook, sorted_balance_sheets, offset=offset)
    else:
        logger.warning("No balance sheet data found in results")
    
    # Fill P&L sheet
    sorted_profit_loss = get_sorted_profit_loss_statements(results)
    
    if sorted_profit_loss:
        fill_vysledovka_sheet(workbook, sorted_profit_loss, offset=offset)
    else:
        logger.warning("No P&L data found in results")
    
    # Fill quality report
    try:
        from src.services.quality import validate_interstatement
        inter_issues = validate_interstatement(results, tolerance)
        fill_quality_report_sheet(workbook, results, inter_issues, tolerance)
    except Exception as e:
        logger.error(f"Failed to fill quality report: {e}", exc_info=True)
    
    # Save to BytesIO buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    logger.info(f"Data landing export completed, buffer size: {buffer.getbuffer().nbytes/1024:.1f}KB")
    return buffer

